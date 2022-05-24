import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_price_per_supplier = {}
product_inventory_under_10 = {}

for item_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(item_row, 4).value
    inventory = int(product_list.cell(item_row, 2).value)
    price = product_list.cell(item_row, 3).value
    product_number = int(product_list.cell(item_row, 1).value)
    inventory_price = product_list.cell(item_row, 5) #to edit 5th column in the sheet

    # calculate total products of each company
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] +=1
    else:
        products_per_supplier[supplier_name] = 1

    # calculate total products of each company
    if supplier_name in total_price_per_supplier:
        total_price_per_supplier[supplier_name] += (inventory * price)
    else:
        total_price_per_supplier[supplier_name] = inventory * price


    #Products with inventory value less than 10
    if inventory < 10:
        product_inventory_under_10[product_number] = inventory


    #add value for total inventory price
    inventory_price.value = inventory * price


print(products_per_supplier)
print(total_price_per_supplier)
print(product_inventory_under_10)

inv_file.save("inventory_updated.xlsx")
